"""輸出成信裡那種寬表:欄=期間×銀行(中信/兆豐/國泰/富邦/玉山),列=指標。
指標對齊信件:Trading_CP+NCD+BA、Trading_GB … 並延伸 OCI/AC 各主要債種。
單位:億元。兆豐債種先天不揭露 → 留空。
"""
import re, pdfplumber
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import extract3 as E

CACHE=Path("pdf_cache")
# 信件欄位順序
BANKS=[("5841","中信"),("5843","兆豐"),("5835","國泰"),("5836","富邦"),("5847","玉山")]
PERIODS=[(roc,mth) for roc in range(109,114) for mth in ("02","04")]  # 2020H1..2024H2
def plabel(roc,mth): return f"{1911+roc}{'H1' if mth=='02' else 'H2'}"

def metrics_for(items3):
    """items3 = {'Trading':dict,'OCI':dict,'AC':dict} → 回傳各指標(億)。"""
    T,O,A = items3["Trading"], items3["OCI"], items3["AC"]
    g=lambda d,*k: sum(d.get(x,0) for x in k)/1e5
    return {
        "Trading_CP+NCD+BA": g(T,"商業本票","可轉讓定期存單","國庫券"),
        "Trading_GB":        g(T,"政府公債"),
        "Trading_公司債":     g(T,"公司債"),
        "Trading_金融債":     g(T,"金融債券"),
        "OCI_GB":            g(O,"政府公債"),
        "OCI_公司債":         g(O,"公司債"),
        "OCI_金融債":         g(O,"金融債券"),
        "AC_GB":             g(A,"政府公債"),
        "AC_公司債":          g(A,"公司債"),
        "AC_金融債":          g(A,"金融債券"),
    }
METRIC_ORDER=["Trading_CP+NCD+BA","Trading_GB","Trading_公司債","Trading_金融債",
              "OCI_GB","OCI_公司債","OCI_金融債","AC_GB","AC_公司債","AC_金融債"]

def collect():
    data={}   # (plabel, name) -> {metric:值}  ;  val None=缺檔/無資料
    for roc,mth in PERIODS:
        lbl=plabel(roc,mth)
        for code,name in BANKS:
            p=CACHE/f"{1911+roc}{mth}_{code}_AI3.pdf"
            if name=="兆豐" or not p.exists() or p.stat().st_size<100000:
                data[(lbl,name)]=None; continue
            t="\n".join((pg.extract_text() or "") for pg in pdfplumber.open(p).pages)
            it3={c:E.parse_class(t,c) for c in ("Trading","OCI","AC")}
            data[(lbl,name)]=metrics_for(it3)
    return data

def build(data, out="銀行債券_寬表.xlsx"):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="債種比較"
    green=PatternFill("solid",fgColor="2E5B4E"); white=Font(bold=True,color="FFFFFF",size=10)
    cen=Alignment(horizontal="center",vertical="center")
    thin=Side(style="thin",color="BBBBBB"); bd=Border(thin,thin,thin,thin)
    plabels=[plabel(r,m) for r,m in PERIODS]; banks=[n for _,n in BANKS]
    nb=len(banks)
    # 表頭兩列
    ws.cell(1,1,"(億元)").font=white; ws.cell(1,1).fill=green
    ws.cell(2,1,"銀行合併").font=white; ws.cell(2,1).fill=green
    for pi,pl in enumerate(plabels):
        c0=2+pi*nb
        ws.merge_cells(start_row=1,start_column=c0,end_row=1,end_column=c0+nb-1)
        cell=ws.cell(1,c0,pl); cell.font=white; cell.fill=green; cell.alignment=cen
        for bi,bn in enumerate(banks):
            cc=ws.cell(2,c0+bi,bn); cc.font=white; cc.fill=green; cc.alignment=cen
    # 資料列
    for ri,metric in enumerate(METRIC_ORDER):
        r=3+ri
        mc=ws.cell(r,1,metric); mc.font=Font(bold=True,size=10)
        for pi,pl in enumerate(plabels):
            for bi,bn in enumerate(banks):
                d=data.get((pl,bn)); c0=2+pi*nb+bi
                v = None if d is None else d.get(metric)
                cell=ws.cell(r,c0, "" if (v is None) else (round(v) if abs(v)>=0.5 else ("" if v==0 else round(v,1))))
                cell.alignment=Alignment(horizontal="right"); cell.border=bd
                cell.number_format="#,##0"
    ws.freeze_panes="B3"
    ws.column_dimensions["A"].width=18
    for c in range(2,2+len(plabels)*nb):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width=6
    wb.save(out); return out

if __name__=="__main__":
    d=collect(); out=build(d)
    print("已輸出:",out)
