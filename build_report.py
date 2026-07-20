"""★ 一鍵主程式:解析五家全期間 → 產出含【原生 Excel 圖表】的報表。
   python3 build_report.py            # 用現有快取
   python3 build_report.py --refresh  # 明年新報表:先自動補抓當期個體檔再產出

輸出 銀行債券_完整報表.xlsx:寬表(全期間) + 資料表 + 兩張原生圖儀表板(版面同參考圖)。
圖表可調參數見下方 CONFIG。
"""
import sys, datetime, pdfplumber
from pathlib import Path
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils as XU
import re, json as _json
import extract3 as E
from extract_megabank import parse_megabank, parse_megabank_fvtpl, parse_megabank_main

# 兆豐年報三分類債種:證券部門變動表對不上、標準彙總逐年格式異、附錄明細表旋轉多行。
# 改用「重要會計項目明細表」附錄以座標+推理配對抽出、每類對帳零誤差的驗證值(見 megabank_override.json)。
# 半年報無此附錄→不在 override→N/A。產生方式見 gen 腳本;未來新期需重跑補入(或接 LLM 自動化)。
try: MEGA_OVERRIDE=_json.load(open("megabank_override.json"))["data"]
except (FileNotFoundError, KeyError): MEGA_OVERRIDE={}

def robust_bs_ac(t):
    """兆豐資產負債表 AC 總額(取所有『按攤銷後成本衡量之債務工具投資 [六(X)] 數字』的最大值,
    避開附註引用等小數字)。回傳億元或 None。"""
    vals=[]
    for m in re.finditer(r"按攤銷後成本衡量之債務工具投資(?:\s*六\([一二三四五六七八九十]+\))?\s*\$?\s*([\d,]{6,})", t):
        vals.append(int(m.group(1).replace(",",""))/1e5)
    return max(vals) if vals else None

def bs_ac_esun(t):
    """玉山 BS『按攤銷後成本衡量之債務工具投資[(附註…)] 數字』行最大金額(千元→億)。
    早期(2020H2/2021H1)文字層亂碼→讀不到乾淨大數→None,作為該期不可靠的訊號。"""
    best=0
    for m in re.finditer(r"按攤銷後成本衡量之債務工具投資(?:（[^）]*）)?\s*([\d,]{7,})", t):
        n=m.group(1)
        if "," in n:                                  # 需含千分位,排除亂碼片段
            best=max(best,int(n.replace(",",""))/1e5)
    return best if best else None

def oci_equity_subtotal(t):
    """FVOCI 權益工具(股票/REITs/受益憑證)小計 = 該類期末公允價值。中信/富邦式;其餘 None。"""
    for m in re.finditer(r"透過其他綜合損益按公允價值衡量\s*之?權益工具", t):
        seg=t[m.end():m.end()+600]
        if "股票" in seg or "受益" in seg or "REIT" in seg:
            sm=re.search(r"(小計|小 計|合計|合 計)\s*\$?\s*([\d,]{4,})", seg)
            if sm: return int(sm.group(2).replace(",",""))
    return None

_FE_A="".join(ch+r"\s*" for ch in "透過其他綜合損益按公允價值衡量之金融資產")
_FE_S="".join(ch+r"\s*" for ch in "股票投資")
_FE_PAT=re.compile(_FE_A+r".{0,80}?"+_FE_S+r"\$?\s*([\d,]{5,})")
def fvoci_equity_level(t):
    """FVOCI 權益(股票)後備:oci_equity_subtotal 抓不到時用(兆豐/國泰/玉山)。回傳仟元或 None。
    層1=公允價值等級表 FVOCI 區塊之『股票投資』(兆豐/國泰);
    層2=(四)明細之單一大額『股票投資』,排除 FVTPL/第三等級/處分 上下文(玉山)。"""
    t2=t.replace("\n"," ")
    m=_FE_PAT.search(t2)
    if m: return int(m.group(1).replace(",",""))
    cand=[]
    for mm in re.finditer(r"股票投資\s*\$?\s*([\d,]{6,})", t2):
        pre=t2[max(0,mm.start()-45):mm.start()]
        if any(k in pre for k in ("透過損益","強制","第三等級","處分","移轉")): continue
        v=int(mm.group(1).replace(",",""))
        if 5_000_000<=v<=200_000_000: cand.append(v)
    return max(cand) if cand else None

# ===================== CONFIG(想調圖就改這裡)=====================
GAP_WIDTH   = 20        # 長條群間距(越小越擠;參考圖約 20~40)
CHART_W, CHART_H = 11, 6.5
START_ROC   = 109       # 起始:民國109 = 2020
SHOW_N      = 6         # 圖表顯示「最近幾期(有資料的)」;寬表仍保留全部
# ================================================================

CACHE=Path("pdf_cache"); OUT="銀行債券_完整報表.xlsx"
BANKS=[("5841","中信"),("5843","兆豐"),("5835","國泰"),("5836","富邦"),("5847","玉山")]
# 自動延伸到「當前民國年」→ 明年後年跑會自動含最新一期,不用改程式重打包
END_ROC=datetime.date.today().year-1911
ALL_PERIODS=[(roc,mth) for roc in range(START_ROC,END_ROC+1) for mth in ("02","04")]
SHOW_PERIODS=[]         # 於 build() 依實際有資料的期間自動決定(取最近 SHOW_N 期)
COLOR={"中信":"4a5e2a","兆豐":"8a8a3a","國泰":"e8c020","富邦":"3a8fd0","玉山":"8bc34a"}
def plabel(roc,mth): return f"{1911+roc}{'H1' if mth=='02' else 'H2'}"

# ---------- 解析 ----------
def parse_all():
    if "--refresh" in sys.argv:
        import resolve
        for roc,mth in ALL_PERIODS:
            for code,_ in BANKS:
                resolve.download(code,roc,mth)
    rec={}
    for roc,mth in ALL_PERIODS:
        lbl=plabel(roc,mth)
        for code,name in BANKS:
            p=CACHE/f"{1911+roc}{mth}_{code}_AI3.pdf"
            if not p.exists() or p.stat().st_size<100000:
                rec[(lbl,name)]=None; continue
            t="\n".join((pg.extract_text() or "") for pg in pdfplumber.open(p).pages)
            if len(t)<2000:      # 無文字層(掃描影像檔,如2020H1國泰/玉山)→ 無資料,非0
                rec[(lbl,name)]=None; continue
            if name=="兆豐":
                # 兆豐三分類:OCI/AC 讀主附註六(四)(五)彙總「毛額」(parse_megabank_main,
                # 詞座標重組+對帳,口徑與其他四家一致);Trading 讀附錄 FVTPL 明細表
                # (parse_megabank_fvtpl)。任一類對帳不過→退回 megabank_override.json
                # 手工對帳值(僅年報有;AC 無評價調整故 override=毛額,退回一致)。全缺→N/A。
                EMPTY={**{k:0.0 for k in E.bond_buckets({})},"股票":0.0}
                mn=parse_megabank_main(p); fv=parse_megabank_fvtpl(p); ov=MEGA_OVERRIDE.get(lbl)
                def _ovc(c): return ({**{k:ov[c].get(k,0.0) for k in E.bond_buckets({})},
                                      "股票":ov[c].get("股票",0.0)} if ov else None)
                def _mk(c):
                    if c=="Trading":
                        if fv and fv.get("_ok"):
                            return {**{k:fv.get(k,0.0) for k in E.bond_buckets({})},"股票":fv.get("股票",0.0)}
                    elif mn["ok"][c]:                     # OCI/AC 主附註對帳過→用毛額
                        return {**E.bond_buckets(mn[c]),"股票":mn["股票"][c]/1e5}
                    return _ovc(c)                        # 後備:override
                parts={c:_mk(c) for c in ("Trading","OCI","AC")}
                if all(v is None for v in parts.values()):
                    rec[(lbl,name)]=None; continue
                r={c:(parts[c] or dict(EMPTY)) for c in ("Trading","OCI","AC")}
                r["_cp"]=(ov.get("_cp",0.0) if ov else 0.0)
                rec[(lbl,name)]=r
                continue
            items={c:E.parse_class(t,c) for c in ("Trading","OCI","AC")}
            if name=="富邦":                                   # 富邦 FVTPL 主附註把國庫券/公司債/公債塞進「其他」→改讀附錄明細表二
                fv=E.parse_fubon_fvtpl(p)
                if fv: items["Trading"]=fv
            r={c:E.bond_buckets(items[c]) for c in ("Trading","OCI","AC")}
            r["_cp"]=items["Trading"].get("商業本票",0)/1e5
            # 股票(權益工具,非債券):FVTPL=股票+受益憑證(已在items);FVOCI=權益工具小計;AC無
            r["Trading"]["股票"]=(items["Trading"].get("股票",0)+items["Trading"].get("受益憑證",0))/1e5
            oe=oci_equity_subtotal(t) or fvoci_equity_level(t); r["OCI"]["股票"]=(oe/1e5) if oe else 0.0
            r["AC"]["股票"]=0.0
            # 玉山早期(2020H2/2021H1)文字層亂碼→分類抓錯表(OCI 抓到股利收入、AC 抓子集)。
            # 守衛:AC 對不上 BS(讀不到乾淨大數 or 差>10%)→整期 N/A,不出誤導數。
            if name=="玉山":
                bs=bs_ac_esun(t); acp=sum(items["AC"].values())/1e5   # 用原始 parse 加總(對得起 BS)
                if not bs or abs(acp-bs)>0.10*bs:
                    rec[(lbl,name)]=None; continue
            rec[(lbl,name)]=r
    return rec

# ---------- 指標 ----------
def mv(b): return b["公債"]+b["公司債"]+b["金融債"]+b["資產基礎"]+b["國庫券"]+b["可轉讓定存單"]+b["其他"]
def tot(r): return sum(mv(r[c]) for c in ("Trading","OCI","AC"))
def clsmv(c): return lambda r: mv(r[c])
def clspct(c): return lambda r: (mv(r[c])/tot(r) if tot(r) else 0)
def typ(k): return lambda r: sum(r[c][k] for c in ("Trading","OCI","AC"))
def credit(r): return sum(r[c]["公司債"]+r[c]["金融債"] for c in ("Trading","OCI","AC"))
def typpct(k): return lambda r: (typ(k)(r)/tot(r) if tot(r) else 0)
otherbond=lambda r: typ("資產基礎")(r)+typ("其他")(r)+typ("國庫券")(r)+typ("可轉讓定存單")(r)

DASH1=[("債券MV合計",tot,0),("Trading MV",clsmv("Trading"),0),("OCI MV",clsmv("OCI"),0),("AC MV",clsmv("AC"),0),
       ("Trading比重",clspct("Trading"),1),("OCI比重",clspct("OCI"),1),("AC比重",clspct("AC"),1)]
DASH2=[("債券MV合計",tot,0),("公債MV",typ("公債"),0),("信用債MV",credit,0),("金融債MV",typ("金融債"),0),
       ("公司債MV",typ("公司債"),0),("其他債MV",otherbond,0),("公債比重",typpct("公債"),1),
       ("信用債比重",lambda r:(credit(r)/tot(r) if tot(r) else 0),1)]

# ---------- 原生圖表(列=銀行、欄=期間、每家一色) ----------
def write_block(ws, top, title, fn, pct, rec):
    ws.cell(top,1,title).font=Font(bold=True); hr=top+1
    ws.cell(hr,1,"銀行").font=Font(bold=True)
    for j,per in enumerate(SHOW_PERIODS): ws.cell(hr,2+j,per).font=Font(bold=True)
    for i,(_,bn) in enumerate(BANKS):
        r=hr+1+i; ws.cell(r,1,bn).font=Font(bold=True)
        for j,per in enumerate(SHOW_PERIODS):
            rb=rec.get((per,bn)); v=None if rb is None else fn(rb)
            c=ws.cell(r,2+j, None if v is None else round(v,4 if pct else 1))
            c.number_format="0%" if pct else "#,##0"
    return hr, hr+len(BANKS)

def add_chart(wsd, hr, last, title, pct, anchor, wsc):
    ch=BarChart(); ch.type="col"; ch.grouping="clustered"; ch.title=title
    ch.height=CHART_H; ch.width=CHART_W; ch.gapWidth=GAP_WIDTH
    data=Reference(wsd,min_col=2,max_col=1+len(SHOW_PERIODS),min_row=hr,max_row=last)
    cats=Reference(wsd,min_col=1,min_row=hr+1,max_row=last)
    ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
    for s in ch.series:                       # 每序列的第 i 點染成第 i 家色
        for i,(_,bn) in enumerate(BANKS):
            dp=DataPoint(idx=i); dp.graphicalProperties=GraphicalProperties(solidFill=COLOR[bn])
            s.data_points.append(dp)
    ch.legend=None
    ch.x_axis.title="銀行"; ch.y_axis.title=("占比" if pct else "億元")
    ch.x_axis.delete=False; ch.y_axis.delete=False
    if pct: ch.y_axis.numFmt="0%"
    wsc.add_chart(ch, anchor)

# ---------- 寬表 ----------
def wide_sheet(ws, rec):
    green=PatternFill("solid",fgColor="2E5B4E"); white=Font(bold=True,color="FFFFFF",size=10)
    cen=Alignment(horizontal="center"); thin=Side(style="thin",color="CCCCCC"); bd=Border(thin,thin,thin,thin)
    METRICS=["Trading_CP+NCD+BA","Trading_GB","Trading_公司債","Trading_金融債","OCI_GB","OCI_公司債","OCI_金融債","AC_GB","AC_公司債","AC_金融債"]
    def metric(rbk,m):
        if rbk is None: return None
        cls,tp=m.split("_",1); b=rbk[cls]
        if tp=="CP+NCD+BA": return rbk["_cp"]+b["國庫券"]+b["可轉讓定存單"]
        return {"GB":b["公債"],"公司債":b["公司債"],"金融債":b["金融債"]}.get(tp)
    plabels=[plabel(r,m) for r,m in ALL_PERIODS]; banks=[n for _,n in BANKS]; nb=len(banks)
    ws.cell(1,1,"(億元)").font=white; ws.cell(1,1).fill=green
    ws.cell(2,1,"銀行合併").font=white; ws.cell(2,1).fill=green
    for pi,pl in enumerate(plabels):
        c0=2+pi*nb; ws.merge_cells(start_row=1,start_column=c0,end_row=1,end_column=c0+nb-1)
        h=ws.cell(1,c0,pl); h.font=white; h.fill=green; h.alignment=cen
        for bi,bn in enumerate(banks):
            hc=ws.cell(2,c0+bi,bn); hc.font=white; hc.fill=green; hc.alignment=cen
    for ri,mn in enumerate(METRICS):
        r=3+ri; ws.cell(r,1,mn).font=Font(bold=True,size=10)
        for pi,pl in enumerate(plabels):
            for bi,bn in enumerate(banks):
                v=metric(rec.get((pl,bn)),mn); c0=2+pi*nb+bi
                cc=ws.cell(r,c0,"" if v is None else round(v)); cc.alignment=Alignment(horizontal="right"); cc.border=bd; cc.number_format="#,##0"
    ws.freeze_panes="B3"; ws.column_dimensions["A"].width=18
    for c in range(2,2+len(plabels)*nb): ws.column_dimensions[XU.get_column_letter(c)].width=6

# ---------- 組裝 ----------
def build(rec):
    global SHOW_PERIODS
    # 圖表期間 = 最近 SHOW_N 個「至少一家有資料」的期間(自動,不寫死年份)
    have=[plabel(r,m) for r,m in ALL_PERIODS
          if any(rec.get((plabel(r,m),n)) for _,n in BANKS)]
    SHOW_PERIODS=have[-SHOW_N:] if have else [plabel(r,m) for r,m in ALL_PERIODS][-SHOW_N:]
    wb=openpyxl.Workbook(); wide_sheet(wb.active,rec); wb.active.title="寬表"
    for dash,specs in [("圖-按分類",DASH1),("圖-按債種",DASH2)]:
        wsd=wb.create_sheet(f"資料_{dash[-3:]}"); wsc=wb.create_sheet(dash); top=1
        for idx,(title,fn,pct) in enumerate(specs):
            hr,last=write_block(wsd,top,title,fn,pct,rec)
            col="A" if idx%2==0 else "L"; row=1+(idx//2)*14
            add_chart(wsd,hr,last,title,pct,f"{col}{row}",wsc); top=last+3
        wsd.column_dimensions["A"].width=8
    wb.save(OUT)

BONDTYPES=["GB","公司債","金融債","資產基礎","貨幣市場","其他","股票"]
WIDE_METRICS=[f"{c}_{t}" for c in ("Trading","OCI","AC") for t in BONDTYPES]
def wide_metric(r, m):
    if r is None: return None
    cls,tp=m.split("_",1); b=r[cls]
    if tp=="貨幣市場":   # 短天期/貨幣市場:CP(僅FVTPL)+國庫券+可轉讓定存單
        return round((r.get("_cp",0) if cls=="Trading" else 0)+b.get("國庫券",0)+b.get("可轉讓定存單",0))
    return round(b.get({"GB":"公債"}.get(tp,tp),0))

def dump_json(rec, path="data.json"):
    import json
    out={"periods":[plabel(r,m) for r,m in ALL_PERIODS],"banks":[n for _,n in BANKS],
         "wide_metrics":WIDE_METRICS,"data":{},"wide":{}}
    for (lbl,name),r in rec.items():
        k=f"{lbl}|{name}"
        out["data"][k]= None if r is None else {c:{x:r[c][x] for x in ("公債","公司債","金融債","其他")} for c in ("Trading","OCI","AC")}
        out["wide"][k]= None if r is None else {m:wide_metric(r,m) for m in WIDE_METRICS}
    json.dump(out, open(path,"w"), ensure_ascii=False)

if __name__=="__main__":
    rec=parse_all(); build(rec); dump_json(rec)
    print("完成 →",OUT,"+ data.json | 圖表期間:",SHOW_PERIODS)
