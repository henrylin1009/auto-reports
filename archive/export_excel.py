"""輸出成品 Excel:五家 × 三分類 × 債種 × 全期間,每格帶 checksum 驗證標記。
- 明細(tidy):期間/銀行/分類/債種/金額(億)/驗證
- 三張樞紐:Trading / OCI / AC(列=期間, 欄=銀行×債種)
兆豐:債種先天不揭露 → 標記「N/A(未揭露)」。
"""
import re, pdfplumber
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import extract3 as E

CACHE=Path("pdf_cache")
BANKS=[("5841","中信"),("5836","富邦"),("5847","玉山"),("5835","國泰"),("5843","兆豐")]
BONDS=["公債","公司債","金融債","資產基礎","國庫券","可轉讓定存單","其他"]
PERIODS=[(roc,mth) for roc in range(109,114) for mth in ("02","04")]  # 2020H1..2024H2

def label(roc,mth): return f"{1911+roc}{'H1' if mth=='02' else 'H2'}"

def collect():
    rows=[]   # (期間,銀行,分類,{債種:億}, 驗證)
    for roc,mth in PERIODS:
        lbl=label(roc,mth)
        for code,name in BANKS:
            p=CACHE/f"{1911+roc}{mth}_{code}_AI3.pdf"
            if not p.exists() or p.stat().st_size<100000:
                rows.append((lbl,name,None,None,"缺檔")); continue
            t="\n".join((pg.extract_text() or "") for pg in pdfplumber.open(p).pages)
            for cls in ("Trading","OCI","AC"):
                if name=="兆豐":
                    rows.append((lbl,name,cls,None,"N/A(未揭露)")); continue
                it,st,ok=E.checksum(t,cls)
                b=E.bond_buckets(it)
                if not it:            status="✗空"
                elif ok:              status="✓"           # 加總=小計,完整可信
                elif st is None:      status="○無小計"      # 報表無可對小計,大多沒事
                else:                 status="⚠️不符"        # 有小計但對不上,務必查
                rows.append((lbl,name,cls,b,status))
    return rows

def build(rows, out="銀行債券投資_債種分析.xlsx"):
    wb=openpyxl.Workbook()
    hdr=Font(bold=True,color="FFFFFF"); hfill=PatternFill("solid",fgColor="2E5B4E")
    okf=PatternFill("solid",fgColor="E6F4EA"); warnf=PatternFill("solid",fgColor="FDECEA")
    # --- 明細 ---
    ws=wb.active; ws.title="明細"
    cols=["期間","銀行","分類"]+BONDS+["驗證"]
    ws.append(cols)
    for c in range(1,len(cols)+1):
        ws.cell(1,c).font=hdr; ws.cell(1,c).fill=hfill
    for lbl,name,cls,b,status in rows:
        vals=[b.get(k,0) if b else None for k in BONDS] if b else [None]*len(BONDS)
        ws.append([lbl,name,cls or "",*[round(v) if v is not None else "" for v in vals],status])
        r=ws.max_row
        if status=="✓": ws.cell(r,len(cols)).fill=okf
        elif status=="⚠️不符": ws.cell(r,len(cols)).fill=warnf
    for i,w in enumerate([9,7,9]+[9]*len(BONDS)+[10],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes="A2"
    # --- 三張樞紐 ---
    for cls in ("Trading","OCI","AC"):
        ps=wb.create_sheet(cls)
        head=["期間"]+[f"{n}·{bt}" for n,_ in [(x[1],x[0]) for x in BANKS] for bt in ("公債","公司債","金融債")]
        # 簡化:每家只放 公債/公司債/金融債 三主債種
        header=["期間"]
        for _,name in BANKS:
            for bt in ("公債","公司債","金融債"):
                header.append(f"{name}·{bt}")
        ps.append(header)
        for c in range(1,len(header)+1):
            ps.cell(1,c).font=hdr; ps.cell(1,c).fill=hfill
        bylp={}
        for lbl,name,c,b,status in rows:
            if c==cls: bylp[(lbl,name)]=(b,status)
        for lbl in [label(r,m) for r,m in PERIODS]:
            line=[lbl]
            for _,name in BANKS:
                b,status=bylp.get((lbl,name),(None,""))
                for bt in ("公債","公司債","金融債"):
                    line.append(round(b[bt]) if b else "")
            ps.append(line)
        ps.freeze_panes="B2"
    wb.save(out); return out

if __name__=="__main__":
    rows=collect()
    out=build(rows)
    print("已輸出:",out)
    # 摘要
    from collections import Counter
    c=Counter(r[4] for r in rows)
    print("格數統計:", dict(c))
