"""Phase 0:AC 隱藏損失 + OCI/AOCI 未實現損益(4 原始欄位,純 PDF、銀行個體)。
輸出 4 欄(仟元,後續除 1e5 成億):
  ① oci_debt_adj  OCI 債務工具評價調整   附註六(四)     純債券,稅前
  ② oci_eq_adj    OCI 權益工具評價調整   附註六(四)     純股票,稅前
  ③ aoci_fa       其他權益-金融資產未實現 附註六(廿九)   債+股淨額,稅後(真AOCI)
  ④ ac_hidden = ac_fv - ac_book          各家公允價值揭露 純債券,帳外
不做任何相減判斷,原始數字全留,之後在既有欄位上加減即可。
"""
import re, pdfplumber
from pathlib import Path

CACHE = Path("pdf_cache")

def to_num(s):
    if s is None: return None
    s = s.replace(",", "").replace("$", "").strip()
    if s in ("-", ""): return 0
    neg = s.startswith("(") and s.endswith(")"); s = s.strip("()").strip()
    return (-int(s) if neg else int(s)) if s.isdigit() else None

def pages_text(code, roc, mth):
    p = CACHE / f"{1911+roc}{mth}_{code}_AI3.pdf"
    if not p.exists() or p.stat().st_size < 100000: return None
    return [pg.extract_text() or "" for pg in pdfplumber.open(p).pages]

NUM = r"\(?\$?[\d,]{3,}\)?|-"

# ---------- ①② OCI 債務/權益 評價調整(附註六(四)) ----------
def oci_adjustments(text):
    """回傳 (債務評價調整, 權益評價調整),取當期(第一欄)。中信型彙總表(附註六(四))。"""
    def first_adj(seg):
        m = re.search(r"(?:金融資產)?評價調整\s*(" + NUM + r")", seg)
        return to_num(m.group(1)) if m else None
    debt_adj = eq_adj = None
    for d in re.finditer(r"透過其他綜合損益按公允價值衡量\s*之?債務工具", text):
        seg = text[d.end():d.end()+900]
        if "評價調整" in seg and ("小計" in seg or "小 計" in seg):
            debt_adj = first_adj(seg); break
    for e in re.finditer(r"透過其他綜合損益按公允價值衡量\s*之?權益工具", text):
        seg = text[e.end():e.end()+700]
        if "評價調整" in seg and ("股票" in seg or "小計" in seg or "小 計" in seg):
            eq_adj = first_adj(seg); break
    return debt_adj, eq_adj

def oci_debt_adj_recon(text):
    """玉山/國泰型:債務工具投資對帳表(總帳面金額/攤銷後成本/公允價值調整),
    『公允價值調整』首個數 = OCI 債務工具未實現(負=含損)。"""
    for m in re.finditer(r"公允價值調整\s*\(\s*([\d,]{4,})\s*\)", text):
        ctx = text[max(0, m.start()-400):m.start()]
        if "總帳面金額" in ctx and "攤銷後成本" in ctx and "透過其他綜合" in ctx:
            return -to_num(m.group(1))
    return None

def oci_debt_adj(text):
    """① OCI 債務工具未實現(稅前):中信彙總評價調整,否則玉山/國泰對帳表公允價值調整。"""
    v = oci_adjustments(text)[0]
    return v if v is not None else oci_debt_adj_recon(text)

def oci_eq_adj(text):
    """② OCI 權益工具未實現(稅前):目前僅中信彙總表可靠取得。"""
    return oci_adjustments(text)[1]

# ---------- ③ 其他權益-金融資產未實現損益(稅後,債+股) ----------
def _aoci_ctbc(text):
    """中信型:其他權益變動表期末列第2欄(兌換 / 金融資產未實現 / 信用風險 / 合計)。"""
    h = re.search(r"其他權益項目變動情形", text) or re.search(r"3\.\s*其他權益", text)
    if not h: return None
    seg = text[h.end():h.end()+1600]
    m = re.search(r"民國[一二三四五六七八九十]+年(?:[一二三四五六七八九十]+月三十日?|十二月三十一日)\s*\$?\s*"
                  r"(" + NUM + r")\s*(" + NUM + r")\s*(" + NUM + r")\s*(" + NUM + r")", seg)
    return to_num(m.group(2)) if m else None

def _aoci_block(text):
    """國泰/富邦型:其他權益 FVOCI 金融資產變動 block(含『債務工具』『權益工具』兩列)→ 隨後『期末餘額』首個數。"""
    for m in re.finditer(r"權益工具", text):
        ctx = text[max(0, m.start()-260):m.start()]
        if "債務工具" not in ctx or "期初餘額" not in ctx:
            continue                                    # 需在含債務工具的變動 block 內
        tail = text[m.start():m.start()+260]
        em = re.search(r"期末餘額\s*(\()?\s*\$?\s*([\d,]{4,})", tail)
        if em:
            v = to_num(em.group(2))
            return -v if em.group(1) else v          # 前括號 = 負(含損)
    return None

def aoci_financial_asset(text):
    """③ 其他權益-透過OCI金融資產未實現損益(稅後,債+股淨額)。中信廿九型 或 國泰/富邦變動block。"""
    v = _aoci_ctbc(text)
    return v if v is not None else _aoci_block(text)

# ---------- ④ AC 帳面 / AC 公允價值 ----------
def ac_book(text):
    """資產負債表:按攤銷後成本衡量之債務工具投資,取第一欄。"""
    m = re.search(r"按攤銷後成本衡量之債務工具投資[^\d\-]*\$?\s*([\d,]{5,})", text)
    return to_num(m.group(1)) if m else None

def ac_fair(text, book):
    """AC 公允價值。相容兩種揭露格式,取債務工具投資該列後最多 4 個數判斷:
      (A) 帳面/公允雙欄(國泰/富邦『非以公允價值衡量者』表):num0≈BS帳面 且 num1<num0 → 公允=num1
      (B) 公允價值等級表(中信/玉山):num0=各等級(num1..3)加總 且 0.8*book<num0<book → 公允=num0
    兩式都對不上 → None(避免誤把帳面當公允出誤導數;如富邦局部揭露、兆豐座標亂碼)。"""
    if book is None: return None
    # 標題後抓最多 4 個數(允許 $ 、換行、括號負值、'-')
    pat = re.compile(r"按?攤銷後成本衡量之\s*債務工具\s*投資\s*" +
                     r"(?:（附註[^）]*）)?\s*" +               # 略過資產負債表列的 (附註…)
                     r"((?:\s*\$?\s*(?:\(?[\d,]{4,}\)?|-)){1,4})")
    def nums_of(blob):
        return [to_num(x) for x in re.findall(r"\(?[\d,]{4,}\)?|-", blob)]
    cands = []
    for m in re.finditer(pat, text):
        # 排除資產負債表列(標題緊接 '（附註'),它給的是帳面非公允
        if "（附註" in text[m.start():m.start()+30]:
            continue
        ns = [n for n in nums_of(m.group(1)) if n is not None]
        if ns: cands.append(ns)
    def inrange(v): return 0.70*book < v < 1.30*book
    # (B) 先認等級表:合計(num0)=第一~三等級(num1..3)加總 → 公允=合計(可正可負,勿假設含損)
    #     必須先做:等級表的「合計 + 第一等級」很像「帳面 + 公允」,會被(A)誤判
    for ns in cands:
        if len(ns) >= 4 and abs(sum(ns[1:4]) - ns[0]) <= 0.01*max(1, ns[0]) and inrange(ns[0]):
            return ns[0]
    # (A) 帳面/公允雙欄:num0≈帳面,num1=公允(可高於帳面=含益)
    for ns in cands:
        if len(ns) >= 2 and abs(ns[0]-book) <= 0.05*book and ns[1] != ns[0] and inrange(ns[1]):
            return ns[1]
    return None

# ---------- 一列 = 一家一期 ----------
def extract_one(code, name, roc, mth):
    pgs = pages_text(code, roc, mth)
    if pgs is None: return None
    # 兆豐=証券部門變動明細表(直排座標式):③用座標解析、AC帳面用 parse_megabank;
    # ①(稅前OCI債券未實現)兆豐無單一評價調整列→N/A,但兆豐OCI全為債券故③即OCI債券準備(稅後);②無FVOCI股票→N/A
    if name == "兆豐":
        from extract_megabank import parse_megabank, parse_megabank_aoci
        p = CACHE / f"{1911+roc}{mth}_{code}_AI3.pdf"
        mb = parse_megabank(p); aoci = parse_megabank_aoci(p, roc, mth)
        if mb is None and aoci is None: return None
        ac_bk = round(sum(mb["AC"].values()) * 1e5) if mb else None
        return {"oci_debt": None, "oci_eq": None, "aoci": aoci,
                "ac_book": ac_bk, "ac_fv": None, "ac_hidden": None, "ac_hidden_pct": None}
    text = "\n".join(pgs)
    if len(text) < 2000: return None          # 掃描影像檔,無文字層
    book = ac_book(text)
    fv   = ac_fair(text, book)                 # 中信等級表 或 國泰型全額揭露表
    hidden = (fv - book) if (fv is not None and book is not None) else None
    return {
        "oci_debt": oci_debt_adj(text),        # ① 稅前
        "oci_eq":   oci_eq_adj(text),          # ② 稅前
        "aoci":     aoci_financial_asset(text),# ③ 稅後(債+股)
        "ac_book":  book,
        "ac_fv":    fv,
        "ac_hidden":hidden,                    # ④
        "ac_hidden_pct": (hidden/book if hidden is not None and book else None),
    }

BANKS = [("5841","中信"),("5843","兆豐"),("5835","國泰"),("5836","富邦"),("5847","玉山")]
def plabel(roc, mth): return f"{1911+roc}{'H1' if mth=='02' else 'H2'}"

def run_all(start_roc=109):
    import datetime
    end_roc = datetime.date.today().year - 1911
    periods = [(r,m) for r in range(start_roc, end_roc+1) for m in ("02","04")]
    rec = {}
    for roc, mth in periods:
        for code, name in BANKS:
            rec[(plabel(roc,mth), name)] = extract_one(code, name, roc, mth)
    return rec, [plabel(r,m) for r,m in periods]

# 4 原始欄(仟元)→ 億元
FIELDS = [("oci_debt","① OCI債券未實現(稅前)"),("oci_eq","② OCI股票未實現(稅前)"),
          ("aoci","③ AOCI金融資產(稅後)"),("ac_book","AC帳面"),("ac_fv","AC公允價值"),
          ("ac_hidden","④ AC隱藏損失"),("ac_hidden_pct","④ 占AC帳面%")]

def to_excel(rec, periods, path="銀行債券_估值指標_Phase0.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "AC隱藏損失+AOCI"
    hd = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="2E5B4E")
    cen = Alignment(horizontal="center"); rgt = Alignment(horizontal="right")
    thin = Side(style="thin", color="DDDDDD"); bd = Border(thin,thin,thin,thin)
    ws.cell(1,1,"(億元;%為占AC帳面)").font = Font(bold=True)
    ws.cell(1,2,"期間").font=hd; ws.cell(1,2).fill=fill
    ws.cell(1,3,"銀行").font=hd; ws.cell(1,3).fill=fill
    for j,(_,lbl) in enumerate(FIELDS):
        c=ws.cell(1,4+j,lbl); c.font=hd; c.fill=fill; c.alignment=cen
    r=2
    for per in periods:
        for _,name in BANKS:
            d=rec.get((per,name))
            ws.cell(r,2,per); ws.cell(r,3,name).font=Font(bold=True)
            for j,(k,_) in enumerate(FIELDS):
                v=None if d is None else d[k]
                cell=ws.cell(r,4+j)
                if v is None: cell.value="N/A"; cell.font=Font(color="999999")
                elif k=="ac_hidden_pct": cell.value=round(v,4); cell.number_format="0.00%"
                else: cell.value=round(v/1e5,1); cell.number_format="#,##0.0"
                cell.alignment=rgt; cell.border=bd
            r+=1
    ws.freeze_panes="D2"; ws.column_dimensions["A"].width=16
    for col in "BC": ws.column_dimensions[col].width=8
    for j in range(len(FIELDS)): ws.column_dimensions[chr(68+j)].width=17
    wb.save(path); return path

def dump_json(rec, periods, path="phase0.json"):
    """給網頁吃:每格 6 欄(億元;pct 為比率),N/A → null。"""
    import json
    keys = ["oci_debt","oci_eq","aoci","ac_book","ac_fv","ac_hidden","ac_hidden_pct"]
    out = {"periods": periods, "banks": [n for _,n in BANKS], "data": {}}
    for per in periods:
        for _,name in BANKS:
            d = rec.get((per,name)); k = f"{per}|{name}"
            if d is None:
                out["data"][k] = None; continue
            out["data"][k] = {kk: (None if d[kk] is None
                              else (round(d[kk],4) if kk=="ac_hidden_pct" else round(d[kk]/1e5,1)))
                              for kk in keys}
    json.dump(out, open(path,"w"), ensure_ascii=False)
    return path

def coverage(rec, periods):
    print(f"\n{'='*60}\n涵蓋率(非 None 格 / 有檔期數)\n{'='*60}")
    have = {name:sum(1 for per in periods if rec.get((per,name))) for _,name in BANKS}
    for k,lbl in FIELDS:
        if k=="ac_hidden_pct": continue
        line=f"{lbl:<22}"
        for _,name in BANKS:
            n=sum(1 for per in periods if rec.get((per,name)) and rec[(per,name)][k] is not None)
            line+=f" {name}{n:>2}/{have[name]:<2}"
        print(line)

if __name__ == "__main__":
    # 打樣自檢:中信 2024H1 四數必須完全命中
    chk = extract_one("5841","中信",113,"02")
    exp = dict(oci_debt=-6387084, oci_eq=3215398, aoci=-6748819, ac_hidden=-35517131)
    ok = all(chk[k]==v for k,v in exp.items())
    print("中信2024H1自檢:", "✅全對" if ok else f"❌ {chk}")
    rec, periods = run_all()
    out = to_excel(rec, periods)
    jout = dump_json(rec, periods)
    coverage(rec, periods)
    print("\n完成 →", out, "+", jout)
