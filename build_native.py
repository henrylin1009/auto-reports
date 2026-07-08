"""原生 Excel 圖表報表,版面同參考圖:x軸=銀行、每家一串時間長條、每家一色。
作法:類別=銀行、序列=期間,對每個序列做「資料點著色」→ 每根長條依所屬銀行上色。
讀 data.json。用法: python3 build_native.py
"""
import json
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font
import openpyxl.utils as XU

D=json.load(open("data.json"))
PERIODS=D["periods"]; BANKS=D["banks"]; DATA=D["data"]
COLOR={"中信":"4a5e2a","兆豐":"8a8a3a","國泰":"e8c020","富邦":"3a8fd0","玉山":"8bc34a"}

def rb(period,bank): return DATA.get(f"{period}|{bank}")
def mv(b): return b["公債"]+b["公司債"]+b["金融債"]+b["其他"]
def tot(r): return sum(mv(r[c]) for c in ("Trading","OCI","AC"))
def M_clsmv(c): return lambda r: mv(r[c])
def M_clspct(c): return lambda r: (mv(r[c])/tot(r) if tot(r) else 0)
def M_type(k): return lambda r: sum(r[c][k] for c in ("Trading","OCI","AC"))
def M_credit(r): return sum(r[c]["公司債"]+r[c]["金融債"] for c in ("Trading","OCI","AC"))
def M_typepct(k): return lambda r: (M_type(k)(r)/tot(r) if tot(r) else 0)

DASH1=[("債券MV合計",tot,False),("Trading MV",M_clsmv("Trading"),False),("OCI MV",M_clsmv("OCI"),False),
       ("AC MV",M_clsmv("AC"),False),("Trading比重",M_clspct("Trading"),True),("OCI比重",M_clspct("OCI"),True),
       ("AC比重",M_clspct("AC"),True)]
DASH2=[("債券MV合計",tot,False),("公債MV",M_type("公債"),False),("信用債MV",M_credit,False),
       ("金融債MV",M_type("金融債"),False),("公司債MV",M_type("公司債"),False),("其他債MV",M_type("其他"),False),
       ("公債比重",M_typepct("公債"),True),("信用債比重",lambda r:(M_credit(r)/tot(r) if tot(r) else 0),True)]

def write_block(ws, top, title, fn, pct):
    """轉置表:列=銀行、欄=期間(第top列標題;top+1表頭;top+2起5家)。回傳(表頭列, 末列)。"""
    ws.cell(top,1,title).font=Font(bold=True)
    hr=top+1
    ws.cell(hr,1,"銀行").font=Font(bold=True)
    for j,per in enumerate(PERIODS): ws.cell(hr,2+j,per).font=Font(bold=True)
    for i,bn in enumerate(BANKS):
        r=hr+1+i; ws.cell(r,1,bn).font=Font(bold=True)
        for j,per in enumerate(PERIODS):
            rec=rb(per,bn); v=None if rec is None else fn(rec)
            c=ws.cell(r,2+j, None if v is None else round(v,4 if pct else 1))
            c.number_format="0%" if pct else "#,##0"
    return hr, hr+len(BANKS)

def make_chart(wsd, hr, last, title, pct, anchor, wsc):
    ch=BarChart(); ch.type="col"; ch.grouping="clustered"; ch.title=title
    ch.height=6.5; ch.width=11; ch.gapWidth=40; ch.overlap=-10
    # 序列=期間(各欄),類別=銀行(各列)
    data=Reference(wsd, min_col=2, max_col=1+len(PERIODS), min_row=hr, max_row=last)
    cats=Reference(wsd, min_col=1, min_row=hr+1, max_row=last)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    # 每個序列的第 i 點 → 染成第 i 家銀行的顏色(達成「每家一色」)
    for s in ch.series:
        for i,bn in enumerate(BANKS):
            dp=DataPoint(idx=i); dp.graphicalProperties=GraphicalProperties(solidFill=COLOR[bn])
            s.data_points.append(dp)
    ch.legend=None                       # 隱藏圖例(序列是期間,不需)
    if pct: ch.y_axis.numFmt="0%"
    wsc.add_chart(ch, anchor)

def build(out="銀行債券_原生圖表.xlsx"):
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    for dash, specs in [("圖-按分類",DASH1),("圖-按債種",DASH2)]:
        wsd=wb.create_sheet(f"資料_{dash[-3:]}"); wsc=wb.create_sheet(dash)
        top=1
        for idx,(title,fn,pct) in enumerate(specs):
            hr,last=write_block(wsd, top, title, fn, pct)
            col="A" if idx%2==0 else "L"; row=1+(idx//2)*14
            make_chart(wsd, hr, last, title, pct, f"{col}{row}", wsc)
            top=last+3
        wsd.column_dimensions["A"].width=8
    wb.save(out); return out

if __name__=="__main__":
    print("完成 →", build())
